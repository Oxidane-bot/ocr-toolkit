"""
Excel data processor for direct data extraction.

This module provides a specialized processor for Excel files that extracts
structured data directly without requiring OCR or PDF conversion.
"""

import logging
import time
from datetime import datetime
from pathlib import Path

from .base import FileProcessorBase, ProcessingResult


class ExcelDataProcessor(FileProcessorBase):
    """
    Processor for Excel files that extracts data directly.

    This class handles:
    - Direct data extraction from .xlsx and .xls files
    - Multiple worksheet support
    - Conversion to Markdown tables
    - Proper handling of formulas, dates, and cell values
    """

    SUPPORTED_FORMATS = {'.xlsx', '.xls'}

    def __init__(self):
        """Initialize Excel data processor."""
        super().__init__()
        self.logger = logging.getLogger(__name__)

    @classmethod
    def supports_format(cls, file_extension: str) -> bool:
        """
        Check if the given file format is supported.

        Args:
            file_extension: File extension to check (with or without dot)

        Returns:
            True if format is supported, False otherwise
        """
        ext = file_extension.lower()
        if not ext.startswith('.'):
            ext = '.' + ext
        return ext in cls.SUPPORTED_FORMATS

    def get_supported_formats(self) -> list[str]:
        """
        Get list of supported file formats.

        Returns:
            List of supported file extensions
        """
        return sorted(self.SUPPORTED_FORMATS)

    def process(self, file_path: str, **kwargs) -> ProcessingResult:
        """
        Process Excel file by extracting data directly.

        Args:
            file_path: Path to the Excel file
            **kwargs: Additional processing parameters

        Returns:
            ProcessingResult object with processing results
        """
        start_time = time.time()
        result = self._create_result(file_path, 'excel_data', start_time)

        if not self._validate_file(file_path):
            result.error = f'Invalid file: {file_path}'
            result.processing_time = time.time() - start_time
            return result

        try:
            ext = Path(file_path).suffix.lower()

            if not self.supports_format(ext):
                result.error = f'Unsupported file format for Excel processing: {ext}'
                result.processing_time = time.time() - start_time
                return result

            # Import openpyxl here to provide better error message if not installed
            try:
                import openpyxl
            except ImportError as e:
                result.error = 'openpyxl library not installed. Please install it to process Excel files.'
                self.logger.error(f"openpyxl import failed: {e}")
                result.processing_time = time.time() - start_time
                return result

            # Process the Excel file
            content = self._process_excel_file(file_path, openpyxl)
            result.content = content
            result.success = True

            # Count sheets as "pages"
            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            result.pages = len(wb.sheetnames)
            wb.close()

            self.logger.debug(f"Extracted data from {result.pages} sheets in {file_path}")

        except Exception as e:
            return self._handle_exception(e, result)

        result.processing_time = time.time() - start_time
        return result

    def _process_excel_file(self, file_path: str, openpyxl) -> str:
        """
        Process Excel file and extract data as Markdown.

        Args:
            file_path: Path to the Excel file
            openpyxl: openpyxl module reference

        Returns:
            Markdown formatted content with all sheets
        """
        try:
            # Load workbook in read-only mode for better performance
            # data_only=True extracts formula values instead of formulas
            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        except Exception as e:
            error_msg = str(e).lower()
            if 'password' in error_msg or 'encrypted' in error_msg:
                raise ValueError(f"Cannot process password-protected Excel file: {file_path}")
            elif 'invalid' in error_msg or 'corrupt' in error_msg:
                raise ValueError(f"Excel file appears to be corrupted: {file_path}")
            else:
                raise ValueError(f"Failed to open Excel file: {e}")

        markdown_parts = [f"# {Path(file_path).name}"]

        try:
            # Process each sheet
            for sheet_name in wb.sheetnames:
                try:
                    ws = wb[sheet_name]
                    sheet_md = self._sheet_to_markdown(ws, sheet_name)
                    markdown_parts.append(sheet_md)
                except Exception as e:
                    self.logger.warning(f"Failed to process sheet '{sheet_name}': {e}")
                    markdown_parts.append(f"\n## Sheet: {sheet_name}\n\n*(Error processing this sheet: {e})*\n")

        finally:
            wb.close()

        return "\n\n".join(markdown_parts)

    def _sheet_to_markdown(self, ws, sheet_name: str) -> str:
        """
        Convert a worksheet to Markdown table format.

        Args:
            ws: openpyxl worksheet object
            sheet_name: Name of the worksheet

        Returns:
            Markdown formatted table
        """
        markdown_lines = [f"## Sheet: {sheet_name}", ""]

        # Check if sheet is empty
        if ws.max_row is None or ws.max_row == 0:
            markdown_lines.append("*(Empty sheet)*")
            return "\n".join(markdown_lines)

        # Extract all rows
        rows = []
        for row in ws.iter_rows(values_only=True):
            # Format each cell value
            formatted_row = [self._format_cell_value(cell) for cell in row]
            rows.append(formatted_row)

        # Skip if no data
        if not rows:
            markdown_lines.append("*(Empty sheet)*")
            return "\n".join(markdown_lines)

        # Determine actual column count (ignore trailing empty columns)
        max_cols = max(len(row) for row in rows)

        # Build Markdown table
        for i, row in enumerate(rows):
            # Pad row to max columns
            padded_row = row + [''] * (max_cols - len(row))

            # Create table row
            markdown_lines.append('| ' + ' | '.join(padded_row) + ' |')

            # Add header separator after first row
            if i == 0:
                markdown_lines.append('| ' + ' | '.join(['---'] * max_cols) + ' |')

        return "\n".join(markdown_lines)

    def _format_cell_value(self, cell_value) -> str:
        """
        Format cell value for Markdown output.

        Args:
            cell_value: Cell value from openpyxl

        Returns:
            Formatted string representation
        """
        # Handle None/empty cells
        if cell_value is None:
            return ''

        # Handle datetime objects
        if isinstance(cell_value, datetime):
            return cell_value.strftime('%Y-%m-%d %H:%M:%S')

        # Handle numbers - format with reasonable precision
        if isinstance(cell_value, (int, float)):
            # For integers or floats that are whole numbers, show without decimals
            if isinstance(cell_value, int) or cell_value == int(cell_value):
                return str(int(cell_value))
            # For floats, limit to 2 decimal places to avoid floating point noise
            return f"{cell_value:.2f}"

        # Convert everything else to string
        cell_str = str(cell_value)

        # Escape pipe characters in Markdown tables
        cell_str = cell_str.replace('|', '\\|')

        # Truncate very long cell values to prevent unwieldy tables
        if len(cell_str) > 100:
            cell_str = cell_str[:97] + '...'

        return cell_str
